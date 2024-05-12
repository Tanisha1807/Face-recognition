import tkinter as tk
import cv2
import numpy as np
import os
import face_recognition
from openpyxl import load_workbook
from datetime import datetime  # Import datetime module

# Specify the path to the directory containing images
database_path = "C:\\Users\\HP\\Desktop\\Project New\\sample"
# Initialize known faces and names
known_face_encodings = []
known_face_names = []

# Load images and their corresponding names from the database
for file_name in os.listdir(database_path):
    name = os.path.splitext(file_name)[0]  # Extract name from file name
    image = face_recognition.load_image_file(os.path.join(database_path, file_name))
    encoding = face_recognition.face_encodings(image)[0]
    known_face_encodings.append(encoding)
    known_face_names.append(name)

# Initialize GUI
root = tk.Tk()
root.title("Facial Recognition Attendance System")

# Initialize webcam
video_capture = cv2.VideoCapture(0)

# Initialize variables
attendance_marked = False
attendance_name = ""

# Function to mark attendance and update Excel file
def mark_attendance(name):
    global attendance_marked
    global attendance_name
    attendance_marked = True
    attendance_name = name
    attendance_label.config(text=f"Attendance Marked for {attendance_name}")

    # Update Excel file
    update_excel(name)

# Function to update attendance in Excel file
def update_excel(name):
    excel_file = "C:\\Users\\HP\\Desktop\\Project New\\attendance.xlsx"  # Path to the Excel file
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()
    worksheet = workbook.active

    # Find the last row in the Excel sheet
    last_row = worksheet.max_row
    # Write the name and date in the next row
    worksheet.cell(row=last_row + 1, column=1).value = name
    worksheet.cell(row=last_row + 1, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Save the workbook
    workbook.save(excel_file)

# Function to update video feed
def update_feed():
    ret, frame = video_capture.read()

    # Find faces in frame
    face_locations = face_recognition.face_locations(frame)
    face_encodings = face_recognition.face_encodings(frame, face_locations)

    for face_encoding in face_encodings:
        # Check if the face matches any known face
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        # If a match is found, mark attendance
        if True in matches:
            match_index = matches.index(True)
            name = known_face_names[match_index]
            if not attendance_marked:
                mark_attendance(name)

        # Draw rectangle around the face
        top, right, bottom, left = face_locations[0]
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

    if ret:
        frame = cv2.flip(frame, 1)  # Flip the frame horizontally
        frame = cv2.resize(frame, (500, 500))  # Resize for display
        frame = np.rot90(frame, 2)  # Rotate for correct orientation in tkinter
        frame = np.flipud(frame)  # Flip for correct orientation in tkinter
        frame = tk.PhotoImage(data=cv2.imencode('.png', frame)[1].tobytes())
        video_label.configure(image=frame)
        video_label.image = frame
        video_label.after(10, update_feed)

# GUI elements
video_label = tk.Label(root)
video_label.pack(padx=10, pady=10)

attendance_label = tk.Label(root, text="Attendance Not Marked")
attendance_label.pack(pady=10)

update_feed()

root.mainloop()

# Release webcam and destroy windows
video_capture.release()
cv2.destroyAllWindows() 